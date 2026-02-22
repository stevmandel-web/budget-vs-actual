"""
Parses the Mapping Tab Excel file to build GL account → P&L line item lookups.

The mapping tab has two independent mapping sections:
  - Columns B-C: Simple GL names (e.g., "General Marketing" → "Advertising Expense")
  - Columns J-K: Hierarchical GL codes (e.g., "13 03 Technician Labor:14 Technician Payroll Wages" → "BT Wages")

Some J-K entries intentionally map to None (non-P&L items like depreciation).
"""
import openpyxl
from config import (
    MAPPING_SHEET, MAPPING_START_ROW,
    MAPPING_SIMPLE_GL_COL, MAPPING_SIMPLE_PNL_COL,
    MAPPING_HIER_GL_COL, MAPPING_HIER_PNL_COL,
)
from parsers.line_item_mapper import canonical_name


def parse_mapping(filepath):
    """
    Parse the Mapping Tab Excel file.

    Returns dict with keys:
        "simple":       {gl_name: pnl_line_item, ...}   from cols B-C
        "hierarchical": {gl_code: pnl_line_item, ...}   from cols J-K
        "non_pnl":      set of GL names that intentionally have no P&L mapping
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[MAPPING_SHEET]

    simple = {}
    hierarchical = {}
    non_pnl = set()

    for row in range(MAPPING_START_ROW, ws.max_row + 1):
        # Simple mappings: B-C
        gl_simple = ws.cell(row=row, column=MAPPING_SIMPLE_GL_COL).value
        pnl_simple = ws.cell(row=row, column=MAPPING_SIMPLE_PNL_COL).value
        if gl_simple and str(gl_simple).strip():
            gl_key = str(gl_simple).strip()
            if pnl_simple and str(pnl_simple).strip():
                mapped = canonical_name(str(pnl_simple).strip()) or str(pnl_simple).strip()
                simple[gl_key] = mapped
            # If pnl_simple is None/empty, the simple section just doesn't have a mapping

        # Hierarchical mappings: J-K
        gl_hier = ws.cell(row=row, column=MAPPING_HIER_GL_COL).value
        pnl_hier = ws.cell(row=row, column=MAPPING_HIER_PNL_COL).value
        if gl_hier and str(gl_hier).strip():
            gl_key = str(gl_hier).strip()
            if pnl_hier and str(pnl_hier).strip():
                mapped = canonical_name(str(pnl_hier).strip()) or str(pnl_hier).strip()
                hierarchical[gl_key] = mapped
            else:
                # Explicitly mapped to None = non-P&L item
                non_pnl.add(gl_key)

    wb.close()
    return {
        "simple": simple,
        "hierarchical": hierarchical,
        "non_pnl": non_pnl,
    }
