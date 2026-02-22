"""
Parses the budget Excel file (MASTER 2026 Budget vBase_3.xlsx).
Reads WholeCo_P&L, Home_P&L, Clinic_P&L, and state-level sheets.
Converts values from $000s to whole dollars.

Key discovery: WholeCo_P&L does NOT have COGS detail (BT Wages, BCBA Wages, etc.)
It jumps from Total Revenue to Gross Profit. COGS detail is only on Home_P&L and
Clinic_P&L. We derive WholeCo COGS by summing those two segments.
"""
import openpyxl
from config import (
    BUDGET_MULTIPLIER, BUDGET_WHOLECO_MONTH_COLS,
    BUDGET_SEGMENT_MONTH_COLS, BUDGET_STATE_MONTH_COLS,
    BUDGET_HOME_STATES, BUDGET_CLINIC_STATES, DATA_LINE_ITEMS,
)
from parsers.line_item_mapper import canonical_name


def _safe_float(val):
    """Convert a cell value to float, treating None/empty as 0."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _read_budget_by_scanning(ws, month_cols, name_col="B", months=None):
    """
    Read budget data by scanning column B for line item names.
    Returns: {line_item: {month: value_in_dollars, ...}, ...}
    """
    if months is None:
        months = list(month_cols.keys())

    data = {}
    for row in range(1, ws.max_row + 1):
        raw_name = ws[f"{name_col}{row}"].value
        item = canonical_name(raw_name)
        if not item or item not in DATA_LINE_ITEMS:
            continue

        if item not in data:
            data[item] = {}
        for month in months:
            if month not in month_cols:
                continue
            col_letter = month_cols[month]
            raw = _safe_float(ws[f"{col_letter}{row}"].value)
            if month not in data[item] or data[item][month] == 0:
                data[item][month] = raw * BUDGET_MULTIPLIER
            else:
                data[item][month] += raw * BUDGET_MULTIPLIER
    return data


def _merge_segment_data(home_data, clinic_data, months):
    """
    Merge Home and Clinic segment data to derive WholeCo COGS and other items.
    For items that exist in both, sum them. For items in only one, take that value.
    Returns: {line_item: {month: value, ...}, ...}
    """
    all_items = set(list(home_data.keys()) + list(clinic_data.keys()))
    merged = {}

    for item in all_items:
        merged[item] = {}
        for month in months:
            home_val = home_data.get(item, {}).get(month, 0)
            clinic_val = clinic_data.get(item, {}).get(month, 0)
            merged[item][month] = home_val + clinic_val

    return merged


def parse_budget(filepath, months=None):
    """
    Parse the full budget file.

    Args:
        filepath: Path to the budget Excel file.
        months: List of months to extract (e.g., ["Jan"]). None = all months.

    Returns: dict with keys:
        "wholeco": {line_item: {month: value, ...}, ...}
        "home":    {line_item: {month: value, ...}, ...}
        "clinic":  {line_item: {month: value, ...}, ...}
        "home_states": {state: {line_item: {month: value}, ...}, ...}
        "clinic_states": {state: {line_item: {month: value}, ...}, ...}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}

    if months is None:
        months = list(BUDGET_WHOLECO_MONTH_COLS.keys())

    # ── Parse WholeCo P&L by scanning (not by fixed rows) ──────────────
    if "WholeCo_P&L" in wb.sheetnames:
        ws = wb["WholeCo_P&L"]
        result["wholeco"] = _read_budget_by_scanning(
            ws, BUDGET_WHOLECO_MONTH_COLS, months=months
        )
    else:
        result["wholeco"] = {}

    # ── Parse Home P&L ──────────────────────────────────────────────────
    if "Home_P&L" in wb.sheetnames:
        ws = wb["Home_P&L"]
        result["home"] = _read_budget_by_scanning(
            ws, BUDGET_SEGMENT_MONTH_COLS, months=months
        )
    else:
        result["home"] = {}

    # ── Parse Clinic P&L ────────────────────────────────────────────────
    if "Clinic_P&L" in wb.sheetnames:
        ws = wb["Clinic_P&L"]
        result["clinic"] = _read_budget_by_scanning(
            ws, BUDGET_SEGMENT_MONTH_COLS, months=months
        )
    else:
        result["clinic"] = {}

    # ── Derive WholeCo COGS from segments ───────────────────────────────
    # WholeCo_P&L doesn't have COGS detail (BT Wages, BCBA Wages, etc.)
    # We merge Home + Clinic to fill in the missing items.
    merged = _merge_segment_data(result["home"], result["clinic"], months)
    for item, month_data in merged.items():
        if item not in result["wholeco"]:
            # Item only exists in segments (COGS detail), add to WholeCo
            result["wholeco"][item] = month_data
        else:
            # Item exists in both — prefer WholeCo value (it's the consolidated)
            # But if WholeCo has 0 and segment has data, use segment
            for m in months:
                if result["wholeco"][item].get(m, 0) == 0 and month_data.get(m, 0) != 0:
                    result["wholeco"][item][m] = month_data[m]

    # ── Parse Home state sheets ─────────────────────────────────────────
    result["home_states"] = {}
    for state in BUDGET_HOME_STATES:
        if state in wb.sheetnames:
            ws = wb[state]
            result["home_states"][state] = _read_budget_by_scanning(
                ws, BUDGET_STATE_MONTH_COLS, months=months
            )

    # ── Parse Clinic state sheets ───────────────────────────────────────
    result["clinic_states"] = {}
    for state in BUDGET_CLINIC_STATES:
        if state in wb.sheetnames:
            ws = wb[state]
            state_clean = state.rstrip("_")
            result["clinic_states"][state_clean] = _read_budget_by_scanning(
                ws, BUDGET_SEGMENT_MONTH_COLS, months=months
            )

    # ── Parse working days from WholeCo_P&L row 4 ───────────────────────
    # Always extract ALL months' working days (needed for Rev/Working Day
    # calculations across current and prior months)
    result["working_days"] = {}
    if "WholeCo_P&L" in wb.sheetnames:
        ws = wb["WholeCo_P&L"]
        for m, col_letter in BUDGET_WHOLECO_MONTH_COLS.items():
            val = ws[f"{col_letter}4"].value
            if val is not None:
                try:
                    result["working_days"][m] = int(float(val))
                except (ValueError, TypeError):
                    pass

    # ── Parse Corporate sheet for management budget ────────────────────
    # The Corporate sheet is a staff roster, but total corporate expenses
    # are already captured in WholeCo_P&L. We flag that it's available.
    result["has_corporate_detail"] = "Corporate" in wb.sheetnames

    wb.close()
    return result
