"""
Parses the actuals Excel file (e.g., January Financials.xlsx).
Reads Combined (or Totalv2), Home, Clinics, and StatebyState/monthly sheets.
Values are already in whole dollars.

Key discovery: The Combined/Home/Clinics sheets have:
  - Row 2: datetime headers (e.g., datetime(2026, 1, 1)) for month columns,
            plus "Budget" and "Variance" text headers
  - Row 3+: Line item names in column B, values in data columns
"""
import openpyxl
from datetime import datetime
from config import MONTHS, DATA_LINE_ITEMS, ACTUALS_STATES
from parsers.line_item_mapper import canonical_name

# Month number to abbreviation
MONTH_NUM_TO_ABBR = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}

# Known individual clinic tabs in actuals
CLINIC_TABS = [
    "AZ-Mesa", "AZ-Phoenix", "AZ-Thunderbird", "AZ-Scottsdale",
    "AZ-Glendale", "AZ-Tucson",
    "NC-Charlotte", "NC-Raeford", "NC-Pinehurst", "NC-WinstonSalem",
    "GA-Savannah",
    "UT-Jordan", "UT-Provo",
    "NM-Albuquerque",
    "Killeen-Clinic",
]


def _safe_float(val):
    """Convert a cell value to float, treating None/empty as 0."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _detect_columns(ws, target_month="Jan"):
    """
    Detect the target month column and budget column.
    Handles both text headers and datetime headers.
    Scans rows 2, 3, and 4 for headers.

    Returns: (target_col_letter, budget_col_letter)
    """
    target_col = None
    budget_col = None

    for hrow in [2, 3, 4]:
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=hrow, column=col).value
            if val is None:
                continue
            col_letter = openpyxl.utils.get_column_letter(col)

            # Check for datetime header
            if isinstance(val, datetime):
                month_abbr = MONTH_NUM_TO_ABBR.get(val.month)
                if month_abbr and month_abbr.lower() == target_month.lower():
                    target_col = col_letter

            # Check for text header
            elif isinstance(val, str):
                val_str = val.strip()
                if val_str.lower().startswith(target_month.lower()):
                    target_col = col_letter
                elif val_str.lower() in ["budget", "bud", "bud."]:
                    budget_col = col_letter

        # If we found the target column, stop scanning header rows
        if target_col:
            break

    return target_col, budget_col


def _detect_all_month_columns(ws):
    """
    Detect ALL month columns (not just target month).
    Returns: {month_abbr: col_letter, ...} for all months found.
    """
    month_cols = {}
    for hrow in [2, 3, 4]:
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=hrow, column=col).value
            if val is None:
                continue
            col_letter = openpyxl.utils.get_column_letter(col)

            if isinstance(val, datetime):
                month_abbr = MONTH_NUM_TO_ABBR.get(val.month)
                if month_abbr:
                    month_cols[month_abbr] = col_letter
            elif isinstance(val, str):
                val_str = val.strip()
                for m_abbr in MONTHS:
                    if val_str.lower().startswith(m_abbr.lower()):
                        month_cols[m_abbr] = col_letter
                        break
        if month_cols:
            break
    return month_cols


def _scan_pnl_sheet(ws, value_col, name_col="B"):
    """
    Read P&L data from an actuals sheet by scanning for line item names.
    Returns: {canonical_name: value, ...}
    """
    data = {}
    for row in range(1, ws.max_row + 1):
        raw_name = ws[f"{name_col}{row}"].value
        item = canonical_name(raw_name)
        if not item or item not in DATA_LINE_ITEMS:
            continue

        val = _safe_float(ws[f"{value_col}{row}"].value)

        if item not in data or data[item] == 0:
            data[item] = val
        else:
            # Accumulate if same item appears multiple times
            data[item] += val

    return data


def _scan_pnl_sheet_all_months(ws, month_cols, name_col="B"):
    """
    Read P&L data for ALL available months from a sheet.
    Returns: {month_abbr: {line_item: value, ...}, ...}
    """
    result = {}
    for month_abbr, col_letter in month_cols.items():
        result[month_abbr] = _scan_pnl_sheet(ws, col_letter, name_col)
    return result


def _scan_state_sheet(ws, state_columns, name_col="B"):
    """
    Read the StatebyState or monthly sheet where each column is a state.
    state_columns: {state_name: col_letter, ...}
    Returns: {state: {line_item: value, ...}, ...}
    """
    result = {}
    for state, col_letter in state_columns.items():
        state_data = {}
        for row in range(1, ws.max_row + 1):
            raw_name = ws[f"{name_col}{row}"].value
            item = canonical_name(raw_name)
            if not item or item not in DATA_LINE_ITEMS:
                continue
            val = _safe_float(ws[f"{col_letter}{row}"].value)
            if item not in state_data or state_data[item] == 0:
                state_data[item] = val
            else:
                state_data[item] += val
        result[state] = state_data
    return result


def _detect_state_columns(ws):
    """
    Detect state columns in the StatebyState/January sheet.
    Tries rows 2, 3, 4 for headers.
    Returns: {state: col_letter, ...} and total_col letter.
    """
    state_cols = {}
    total_col = None

    for header_row in [2, 3, 4]:
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            if val is None:
                continue
            val_str = str(val).strip()
            col_letter = openpyxl.utils.get_column_letter(col)

            if val_str.upper() == "TOTAL":
                total_col = col_letter
            else:
                for s in ACTUALS_STATES:
                    if s.upper() == val_str.upper():
                        state_cols[s] = col_letter
                        break
        if state_cols:
            break

    return state_cols, total_col


def parse_actuals(filepath, target_month="Jan"):
    """
    Parse the actuals file for a specific month.

    Args:
        filepath: Path to the actuals Excel file.
        target_month: Month to extract (e.g., "Jan").

    Returns: dict with keys:
        "wholeco": {line_item: value, ...}
        "home":    {line_item: value, ...}
        "clinic":  {line_item: value, ...}
        "states":  {state: {line_item: value, ...}, ...}
        "mgmt":    {line_item: value, ...}
        "clinics_detail": {clinic_name: {line_item: value, ...}, ...}
        "historical": {month: {line_item: value, ...}, ...}
        "target_month": the month extracted
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {"target_month": target_month}

    # ── Detect columns using the Combined sheet first (has latest data) ──
    target_col = None
    budget_col = None

    # Priority order: Combined has Oct-Jan; Totalv2 only has Jun-Nov
    for sheet_name in ["Combined", "Home", "Clinics", "Totalv2"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        tc, bc = _detect_columns(ws, target_month)
        if tc:
            target_col = tc
            budget_col = bc
            break

    if not target_col:
        print(f"  WARNING: Could not find column for {target_month}. Defaulting to F.")
        target_col = "F"

    # ── Parse WholeCo ───────────────────────────────────────────────────
    for sheet_name in ["Combined", "Totalv2"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            tc, _ = _detect_columns(ws, target_month)
            col = tc or target_col
            result["wholeco"] = _scan_pnl_sheet(ws, col)
            if result["wholeco"].get("Total Revenue", 0) != 0:
                break
    else:
        result["wholeco"] = {}

    # ── Parse Home ──────────────────────────────────────────────────────
    if "Home" in wb.sheetnames:
        ws = wb["Home"]
        tc, _ = _detect_columns(ws, target_month)
        col = tc or target_col
        result["home"] = _scan_pnl_sheet(ws, col)
    else:
        result["home"] = {}

    # ── Parse Clinics ───────────────────────────────────────────────────
    if "Clinics" in wb.sheetnames:
        ws = wb["Clinics"]
        tc, _ = _detect_columns(ws, target_month)
        col = tc or target_col
        result["clinic"] = _scan_pnl_sheet(ws, col)
    else:
        result["clinic"] = {}

    # ── Parse StatebyState ──────────────────────────────────────────────
    for sheet_name in ["StatebyState", "January", target_month]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            state_cols, total_col = _detect_state_columns(ws)
            if state_cols:
                result["states"] = _scan_state_sheet(ws, state_cols)
                break
    else:
        result["states"] = {}

    # ── Parse MGMT (Management cost center) ─────────────────────────────
    if "MGMT" in wb.sheetnames:
        ws = wb["MGMT"]
        tc, _ = _detect_columns(ws, target_month)
        col = tc or target_col
        result["mgmt"] = _scan_pnl_sheet(ws, col)
    else:
        result["mgmt"] = {}

    # ── Parse individual clinic tabs ────────────────────────────────────
    result["clinics_detail"] = {}
    for clinic_name in CLINIC_TABS:
        if clinic_name in wb.sheetnames:
            ws = wb[clinic_name]
            tc, _ = _detect_columns(ws, target_month)
            col = tc or target_col
            clinic_data = _scan_pnl_sheet(ws, col)
            # Only include if clinic has meaningful data
            if clinic_data.get("Total Revenue", 0) != 0 or clinic_data.get("BT Revenue", 0) != 0:
                result["clinics_detail"][clinic_name] = clinic_data

    # ── Parse historical months from Combined sheet ─────────────────────
    result["historical"] = {}
    for sheet_name in ["Combined", "Totalv2"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_month_cols = _detect_all_month_columns(ws)
            if all_month_cols:
                result["historical"] = _scan_pnl_sheet_all_months(ws, all_month_cols)
                break

    # ── Parse historical state-by-state (Oct, Nov, Dec tabs) ────────────
    result["historical_states"] = {}
    for month_name in ["October", "November", "December", "January",
                        "Oct", "Nov", "Dec", "Jan"]:
        if month_name in wb.sheetnames:
            ws = wb[month_name]
            state_cols, _ = _detect_state_columns(ws)
            if state_cols:
                # Map to standard month abbreviation
                m_abbr = month_name[:3].capitalize()
                if m_abbr not in result["historical_states"]:
                    result["historical_states"][m_abbr] = _scan_state_sheet(ws, state_cols)

    wb.close()
    return result
