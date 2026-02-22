"""
Excel formatting helpers: styles, colors, number formats.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers


# ── Colors ──────────────────────────────────────────────────────────────────
GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
RED_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
GREEN_FONT = Font(color="1B5E20")
RED_FONT = Font(color="C62828")
HEADER_FILL = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SECTION_FILL = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11, color="1A237E")
SUBTOTAL_FONT = Font(bold=True, size=10)
TOTAL_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
ITEM_FONT = Font(size=10)

# Insight severities
SEVERITY_FILLS = {
    "critical": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
    "warning": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
    "positive": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    "info": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
}

SEVERITY_FONTS = {
    "critical": Font(color="B71C1C", bold=True),
    "warning": Font(color="F57F17", bold=True),
    "positive": Font(color="1B5E20", bold=True),
    "info": Font(color="0D47A1", bold=True),
}

# ── Borders ─────────────────────────────────────────────────────────────────
THIN_BORDER = Border(
    bottom=Side(style="thin", color="BDBDBD"),
)
THICK_BORDER = Border(
    top=Side(style="medium", color="424242"),
    bottom=Side(style="medium", color="424242"),
)
BOTTOM_THICK = Border(
    bottom=Side(style="medium", color="424242"),
)

# ── Number formats ──────────────────────────────────────────────────────────
DOLLAR_FMT = '#,##0'
DOLLAR_NEG_FMT = '#,##0;(#,##0)'
DOLLAR_VAR_FMT = '#,##0;(#,##0)'
PCT_FMT = '0.0%'

# ── Alignment ───────────────────────────────────────────────────────────────
RIGHT_ALIGN = Alignment(horizontal="right")
LEFT_ALIGN = Alignment(horizontal="left")
CENTER_ALIGN = Alignment(horizontal="center")
INDENT_ALIGN = Alignment(horizontal="left", indent=2)
