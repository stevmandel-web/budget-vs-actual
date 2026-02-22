"""
Builds the formatted Budget vs Actual output Excel workbook.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from output.formatters import (
    GREEN_FILL, RED_FILL, GREEN_FONT, RED_FONT,
    HEADER_FILL, HEADER_FONT, SECTION_FILL, SECTION_FONT,
    SUBTOTAL_FONT, TOTAL_FONT, TOTAL_FILL, ITEM_FONT,
    SEVERITY_FILLS, SEVERITY_FONTS,
    THIN_BORDER, THICK_BORDER, BOTTOM_THICK,
    DOLLAR_FMT, DOLLAR_VAR_FMT, PCT_FMT,
    RIGHT_ALIGN, LEFT_ALIGN, CENTER_ALIGN, INDENT_ALIGN,
)


def _apply_variance_color(cell, favorable):
    """Apply green/red formatting based on favorable/unfavorable."""
    if favorable is None:
        return
    if favorable:
        cell.font = GREEN_FONT
        cell.fill = GREEN_FILL
    else:
        cell.font = RED_FONT
        cell.fill = RED_FILL


def _write_header_row(ws, row, headers, col_start=1):
    """Write a styled header row."""
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN if i > 0 else LEFT_ALIGN


def _format_row(ws, row, row_type, data_col_start, data_col_end, skip_col=None, skip_cols=None):
    """Apply formatting based on row type (header, item, subtotal, total).
    skip_col: optional single column number to skip (backward compat).
    skip_cols: optional list/set of column numbers to skip (e.g., separator columns)."""
    # Merge skip_col and skip_cols into a single set
    _skip = set()
    if skip_col is not None:
        _skip.add(skip_col)
    if skip_cols:
        _skip.update(skip_cols)

    label_cell = ws.cell(row=row, column=1)

    if row_type == "header":
        label_cell.font = SECTION_FONT
        label_cell.fill = SECTION_FILL
        for col in range(1, data_col_end + 1):
            if col in _skip:
                continue
            ws.cell(row=row, column=col).fill = SECTION_FILL
    elif row_type == "total":
        label_cell.font = TOTAL_FONT
        label_cell.fill = TOTAL_FILL
        for col in range(1, data_col_end + 1):
            if col in _skip:
                continue
            c = ws.cell(row=row, column=col)
            c.font = TOTAL_FONT
            c.fill = TOTAL_FILL
            c.border = THICK_BORDER
    elif row_type == "subtotal":
        label_cell.font = SUBTOTAL_FONT
        for col in range(data_col_start, data_col_end + 1):
            if col in _skip:
                continue
            ws.cell(row=row, column=col).font = SUBTOTAL_FONT
            ws.cell(row=row, column=col).border = THIN_BORDER
    elif row_type == "item":
        label_cell.font = ITEM_FONT
        label_cell.alignment = INDENT_ALIGN


def _write_separator_col(ws, col, last_row):
    """Format a column as a thin visual separator between groups."""
    SEPARATOR_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    for r in range(3, last_row + 1):
        ws.cell(row=r, column=col).fill = SEPARATOR_FILL


def write_wholeco_sheet(wb, variance_rows, month, prior_month_actuals=None, working_days=None):
    """
    Write the WholeCo Summary sheet.
    Column layout:
        A: Line Item
        ── ACTUALS GROUP ──
        B: Dec Actual (prior month)
        C: Jan Actual ★ (current, most recent)
        D: MoM $ Change
        E: [separator]
        ── BUDGET VARIANCE GROUP ──
        F: Jan Budget
        G: $ Variance
        H: % Variance

    If no prior month data: A | Jan Actual ★ | [sep] | Jan Budget | $ Var | % Var
    """
    ws = wb.create_sheet("WholeCo Summary")
    has_prior = prior_month_actuals is not None and len(prior_month_actuals) > 0

    # Determine prior month name
    from config import MONTHS
    month_idx = MONTHS.index(month) if month in MONTHS else 0
    prior_month_name = MONTHS[month_idx - 1]  # Python negative indexing wraps correctly

    # Column assignments depend on whether we have prior month data
    if has_prior:
        COL_PRIOR_ACT = 2    # B: Dec Actual
        COL_CURR_ACT = 3     # C: Jan Actual ★
        COL_MOM = 4          # D: MoM $ Change
        COL_SEP = 5          # E: separator
        COL_BUDGET = 6       # F: Budget
        COL_DOLLAR_VAR = 7   # G: $ Variance
        COL_PCT_VAR = 8      # H: % Variance
        last_col = 8
    else:
        COL_PRIOR_ACT = None
        COL_CURR_ACT = 2     # B: Jan Actual ★
        COL_MOM = None
        COL_SEP = 3          # C: separator
        COL_BUDGET = 4       # D: Budget
        COL_DOLLAR_VAR = 5   # E: $ Variance
        COL_PCT_VAR = 6      # F: % Variance
        last_col = 6

    # Title
    ws.cell(row=1, column=1, value=f"Treetop Therapy — Budget vs Actual: {month} 2026")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1A237E")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

    # Group label row (row 2) — identifies Actuals vs Budget sections
    ACTUALS_LABEL_FILL = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
    BUDGET_LABEL_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    ACTUALS_LABEL_FONT = Font(bold=True, size=10, color="1A237E")
    BUDGET_LABEL_FONT = Font(bold=True, size=10, color="E65100")

    if has_prior:
        ws.merge_cells(start_row=2, start_column=COL_PRIOR_ACT, end_row=2, end_column=COL_MOM)
        c = ws.cell(row=2, column=COL_PRIOR_ACT, value="── Actuals ──")
        c.font = ACTUALS_LABEL_FONT
        c.fill = ACTUALS_LABEL_FILL
        c.alignment = CENTER_ALIGN
        for col in range(COL_PRIOR_ACT, COL_MOM + 1):
            ws.cell(row=2, column=col).fill = ACTUALS_LABEL_FILL

        ws.merge_cells(start_row=2, start_column=COL_BUDGET, end_row=2, end_column=COL_PCT_VAR)
        c = ws.cell(row=2, column=COL_BUDGET, value="── Budget vs Actual ──")
        c.font = BUDGET_LABEL_FONT
        c.fill = BUDGET_LABEL_FILL
        c.alignment = CENTER_ALIGN
        for col in range(COL_BUDGET, COL_PCT_VAR + 1):
            ws.cell(row=2, column=col).fill = BUDGET_LABEL_FILL
    else:
        c = ws.cell(row=2, column=COL_CURR_ACT, value="── Actual ──")
        c.font = ACTUALS_LABEL_FONT
        c.fill = ACTUALS_LABEL_FILL
        c.alignment = CENTER_ALIGN

        ws.merge_cells(start_row=2, start_column=COL_BUDGET, end_row=2, end_column=COL_PCT_VAR)
        c = ws.cell(row=2, column=COL_BUDGET, value="── Budget vs Actual ──")
        c.font = BUDGET_LABEL_FONT
        c.fill = BUDGET_LABEL_FILL
        c.alignment = CENTER_ALIGN
        for col in range(COL_BUDGET, COL_PCT_VAR + 1):
            ws.cell(row=2, column=col).fill = BUDGET_LABEL_FILL

    # Headers row (row 3)
    headers_map = {1: "Line Item"}
    if has_prior:
        headers_map[COL_PRIOR_ACT] = f"{prior_month_name} Actual"
        headers_map[COL_CURR_ACT] = f"{month} Actual ★"
        headers_map[COL_MOM] = "MoM $ Chg"
        headers_map[COL_SEP] = ""
        headers_map[COL_BUDGET] = f"{month} Budget"
        headers_map[COL_DOLLAR_VAR] = "$ Variance"
        headers_map[COL_PCT_VAR] = "% Variance"
    else:
        headers_map[COL_CURR_ACT] = f"{month} Actual ★"
        headers_map[COL_SEP] = ""
        headers_map[COL_BUDGET] = f"{month} Budget"
        headers_map[COL_DOLLAR_VAR] = "$ Variance"
        headers_map[COL_PCT_VAR] = "% Variance"

    for col_num, header_text in headers_map.items():
        cell = ws.cell(row=3, column=col_num, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN if col_num > 1 else LEFT_ALIGN
    # Ensure separator header cell is also styled
    ws.cell(row=3, column=COL_SEP).fill = HEADER_FILL

    # Data rows
    row_num = 4

    # Insert Revenue per Working Day after Total Revenue
    rev_per_day_inserted = False

    # Pre-extract Total Revenue values for Rev/Working Day calculation
    total_rev_actual = 0
    total_rev_budget = 0
    for tr in variance_rows:
        if tr["label"] == "Total Revenue":
            total_rev_actual = tr["actual"] or 0
            total_rev_budget = tr["budget"] or 0
            break

    for vr in variance_rows:
        label = vr["label"]
        row_type = vr["row_type"]
        is_pct_row = row_type == "pct_row"

        if row_type == "blank":
            # After Total Revenue blank, insert Rev/Working Day
            if not rev_per_day_inserted and working_days and working_days.get(month):
                prev_label = None
                for prev_vr in variance_rows:
                    if prev_vr is vr:
                        break
                    if prev_vr["label"]:
                        prev_label = prev_vr["label"]
                if prev_label == "Total Revenue":
                    wd = working_days[month]
                    act_rpd = total_rev_actual / wd if wd else 0
                    bud_rpd = total_rev_budget / wd if wd else 0
                    rpd_var = act_rpd - bud_rpd

                    ws.cell(row=row_num, column=1, value="Revenue per Working Day")
                    ws.cell(row=row_num, column=1).font = ITEM_FONT
                    ws.cell(row=row_num, column=1).alignment = INDENT_ALIGN

                    # Prior month Rev/Working Day (if historical data available)
                    if has_prior:
                        prior_rev = prior_month_actuals.get("Total Revenue", 0)
                        # Try to get prior month working days
                        prior_wd = working_days.get(prior_month_name)
                        if prior_rev and prior_wd:
                            c = ws.cell(row=row_num, column=COL_PRIOR_ACT, value=prior_rev / prior_wd)
                            c.number_format = DOLLAR_FMT
                            c.alignment = RIGHT_ALIGN

                    # Current month Actual Rev/Day
                    c = ws.cell(row=row_num, column=COL_CURR_ACT, value=act_rpd)
                    c.number_format = DOLLAR_FMT
                    c.alignment = RIGHT_ALIGN

                    # MoM for Rev/Day
                    if has_prior:
                        prior_rev = prior_month_actuals.get("Total Revenue", 0)
                        prior_wd = working_days.get(prior_month_name)
                        if prior_rev and prior_wd:
                            prior_rpd = prior_rev / prior_wd
                            mom_rpd = act_rpd - prior_rpd
                            c = ws.cell(row=row_num, column=COL_MOM, value=mom_rpd)
                            c.number_format = DOLLAR_VAR_FMT
                            c.alignment = RIGHT_ALIGN
                            _apply_variance_color(c, mom_rpd >= 0)

                    # Budget Rev/Day
                    c = ws.cell(row=row_num, column=COL_BUDGET, value=bud_rpd)
                    c.number_format = DOLLAR_FMT
                    c.alignment = RIGHT_ALIGN

                    # Variance
                    c = ws.cell(row=row_num, column=COL_DOLLAR_VAR, value=rpd_var)
                    c.number_format = DOLLAR_VAR_FMT
                    c.alignment = RIGHT_ALIGN
                    _apply_variance_color(c, rpd_var >= 0)

                    # Working days note
                    row_num += 1
                    wd_note = f"Working Days: {wd}"
                    if has_prior:
                        prior_wd = working_days.get(prior_month_name)
                        if prior_wd:
                            wd_note = f"Working Days: {month}={wd}, {prior_month_name}={prior_wd}"
                    ws.cell(row=row_num, column=1, value=wd_note)
                    ws.cell(row=row_num, column=1).font = Font(size=9, italic=True, color="666666")
                    rev_per_day_inserted = True
                    row_num += 1

            row_num += 1
            continue

        # Label
        ws.cell(row=row_num, column=1, value=label)

        if row_type not in ("header",):
            num_fmt = PCT_FMT if is_pct_row else DOLLAR_FMT
            var_fmt = PCT_FMT if is_pct_row else DOLLAR_VAR_FMT

            # Prior Month Actual
            if has_prior and COL_PRIOR_ACT:
                if is_pct_row:
                    # Compute pct_row from prior month dollar values
                    prior_rev = prior_month_actuals.get("Total Revenue", 0)
                    prior_pct = None
                    if prior_rev and prior_rev > 0:
                        if label == "Gross Margin, %":
                            prior_pct = prior_month_actuals.get("Gross Profit", 0) / prior_rev
                        elif label == "Billing Expense, %":
                            prior_pct = prior_month_actuals.get("Billing Expense", 0) / prior_rev
                        elif label == "Gross Margin Net Billing, %":
                            prior_pct = prior_month_actuals.get("Gross Profit Net Billing", 0) / prior_rev
                        elif label == "EBITDA, %":
                            prior_pct = prior_month_actuals.get("EBITDA", 0) / prior_rev
                    if prior_pct is not None:
                        c = ws.cell(row=row_num, column=COL_PRIOR_ACT, value=prior_pct)
                        c.number_format = PCT_FMT
                        c.alignment = RIGHT_ALIGN
                else:
                    prior_val = prior_month_actuals.get(label, 0)
                    if prior_val:
                        c = ws.cell(row=row_num, column=COL_PRIOR_ACT, value=prior_val)
                        c.number_format = num_fmt
                        c.alignment = RIGHT_ALIGN

            # Current Month Actual
            if vr["actual"] is not None:
                c = ws.cell(row=row_num, column=COL_CURR_ACT, value=vr["actual"])
                c.number_format = num_fmt
                c.alignment = RIGHT_ALIGN

            # MoM Change
            if has_prior and COL_MOM:
                if is_pct_row:
                    # MoM percentage point change for pct_rows
                    prior_rev = prior_month_actuals.get("Total Revenue", 0)
                    prior_pct = None
                    if prior_rev and prior_rev > 0:
                        if label == "Gross Margin, %":
                            prior_pct = prior_month_actuals.get("Gross Profit", 0) / prior_rev
                        elif label == "Billing Expense, %":
                            prior_pct = prior_month_actuals.get("Billing Expense", 0) / prior_rev
                        elif label == "Gross Margin Net Billing, %":
                            prior_pct = prior_month_actuals.get("Gross Profit Net Billing", 0) / prior_rev
                        elif label == "EBITDA, %":
                            prior_pct = prior_month_actuals.get("EBITDA", 0) / prior_rev
                    if prior_pct is not None and vr["actual"] is not None:
                        mom_pp = vr["actual"] - prior_pct
                        c = ws.cell(row=row_num, column=COL_MOM, value=mom_pp)
                        c.number_format = PCT_FMT
                        c.alignment = RIGHT_ALIGN
                        if vr.get("is_revenue_like"):
                            _apply_variance_color(c, mom_pp >= 0)
                        else:
                            _apply_variance_color(c, mom_pp <= 0)
                else:
                    prior_val = prior_month_actuals.get(label, 0)
                    if vr["actual"] is not None and prior_val:
                        mom_var = vr["actual"] - prior_val
                        c = ws.cell(row=row_num, column=COL_MOM, value=mom_var)
                        c.number_format = DOLLAR_VAR_FMT
                        c.alignment = RIGHT_ALIGN
                        if vr.get("is_revenue_like"):
                            _apply_variance_color(c, mom_var >= 0)
                        else:
                            _apply_variance_color(c, mom_var <= 0)

            # Budget
            if vr["budget"] is not None:
                c = ws.cell(row=row_num, column=COL_BUDGET, value=vr["budget"])
                c.number_format = num_fmt
                c.alignment = RIGHT_ALIGN

            # $ Variance (or pp difference for pct_rows)
            if vr["dollar_var"] is not None:
                c = ws.cell(row=row_num, column=COL_DOLLAR_VAR, value=vr["dollar_var"])
                c.number_format = var_fmt
                c.alignment = RIGHT_ALIGN
                _apply_variance_color(c, vr["favorable"])

            # % Variance (not shown for pct_rows)
            if not is_pct_row and vr["pct_var"] is not None and vr["budget"] != 0:
                c = ws.cell(row=row_num, column=COL_PCT_VAR, value=vr["pct_var"])
                c.number_format = PCT_FMT
                c.alignment = RIGHT_ALIGN
                _apply_variance_color(c, vr["favorable"])

        # Row formatting — skip separator column
        _format_row(ws, row_num, row_type if not is_pct_row else "item", 2, last_col, skip_col=COL_SEP)
        row_num += 1

    # Separator column styling
    _write_separator_col(ws, COL_SEP, row_num - 1)

    # Column widths
    ws.column_dimensions["A"].width = 30
    if has_prior:
        ws.column_dimensions[get_column_letter(COL_PRIOR_ACT)].width = 16
        ws.column_dimensions[get_column_letter(COL_CURR_ACT)].width = 18
        ws.column_dimensions[get_column_letter(COL_MOM)].width = 14
        ws.column_dimensions[get_column_letter(COL_SEP)].width = 2
        ws.column_dimensions[get_column_letter(COL_BUDGET)].width = 16
        ws.column_dimensions[get_column_letter(COL_DOLLAR_VAR)].width = 16
        ws.column_dimensions[get_column_letter(COL_PCT_VAR)].width = 12
    else:
        ws.column_dimensions[get_column_letter(COL_CURR_ACT)].width = 18
        ws.column_dimensions[get_column_letter(COL_SEP)].width = 2
        ws.column_dimensions[get_column_letter(COL_BUDGET)].width = 16
        ws.column_dimensions[get_column_letter(COL_DOLLAR_VAR)].width = 16
        ws.column_dimensions[get_column_letter(COL_PCT_VAR)].width = 12

    # Freeze panes
    ws.freeze_panes = "B4"


def write_segment_sheet(wb, segment_rows, month):
    """
    Write the Home vs Clinic comparison sheet.
    """
    ws = wb.create_sheet("Home vs Clinic")

    # Title
    ws.cell(row=1, column=1, value=f"Treetop Therapy — Home vs Clinic: {month} 2026")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1A237E")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    # Headers — Actual before Budget
    headers = [
        "Line Item",
        "Home Actual", "Home Budget", "Home $ Var", "Home % Var",
        "Clinic Actual", "Clinic Budget", "Clinic $ Var", "Clinic % Var",
    ]
    _write_header_row(ws, 3, headers)

    row_num = 4
    for vr in segment_rows:
        label = vr["label"]
        row_type = vr["row_type"]
        is_pct_row = row_type == "pct_row"

        if row_type == "blank":
            row_num += 1
            continue

        ws.cell(row=row_num, column=1, value=label)

        if row_type not in ("header",):
            num_fmt = PCT_FMT if is_pct_row else DOLLAR_FMT
            var_fmt = PCT_FMT if is_pct_row else DOLLAR_VAR_FMT

            col_offset = 2
            for seg in ["home", "clinic"]:
                # Actual (first)
                act = vr.get(f"{seg}_actual")
                if act is not None:
                    c = ws.cell(row=row_num, column=col_offset, value=act)
                    c.number_format = num_fmt
                    c.alignment = RIGHT_ALIGN

                # Budget (second)
                bud = vr.get(f"{seg}_budget")
                if bud is not None:
                    c = ws.cell(row=row_num, column=col_offset + 1, value=bud)
                    c.number_format = num_fmt
                    c.alignment = RIGHT_ALIGN

                # $ Var
                dvar = vr.get(f"{seg}_dollar_var")
                if dvar is not None:
                    c = ws.cell(row=row_num, column=col_offset + 2, value=dvar)
                    c.number_format = var_fmt
                    c.alignment = RIGHT_ALIGN
                    _apply_variance_color(c, vr.get(f"{seg}_favorable"))

                # % Var (not shown for pct_rows)
                if not is_pct_row:
                    pvar = vr.get(f"{seg}_pct_var")
                    bud_val = vr.get(f"{seg}_budget", 0)
                    if pvar is not None and bud_val != 0:
                        c = ws.cell(row=row_num, column=col_offset + 3, value=pvar)
                        c.number_format = PCT_FMT
                        c.alignment = RIGHT_ALIGN
                        _apply_variance_color(c, vr.get(f"{seg}_favorable"))

                col_offset += 4

        _format_row(ws, row_num, row_type if not is_pct_row else "item", 2, 9)
        row_num += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    for col_letter in ["B", "C", "D", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 16
    for col_letter in ["E", "I"]:
        ws.column_dimensions[col_letter].width = 12

    ws.freeze_panes = "B4"


def write_state_drilldown_sheet(wb, state_variances, month, states):
    """
    Write the State Drill-Down sheet.
    One set of columns (Actual, Budget, $ Var) per state, with gray separator
    columns between states and state group labels in row 2.
    """
    ws = wb.create_sheet("State Drill-Down")

    # Title
    ws.cell(row=1, column=1, value=f"Treetop Therapy — State Drill-Down: {month} 2026")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1A237E")

    from config import PNL_STRUCTURE

    # ── Build column layout with separators ───────────────────────────────
    # Each state gets 3 data cols + 1 separator col (except last state)
    # Track: state_col_starts[state] = first data column for that state
    #        separator_cols = list of separator column numbers
    headers = ["Line Item"]
    state_col_starts = {}
    separator_cols = []
    col = 2  # column A = Line Item, start data at column 2

    for i, state in enumerate(states):
        state_col_starts[state] = col
        headers.append(f"{state} Act")
        headers.append(f"{state} Bud")
        headers.append(f"{state} $ Var")
        col += 3

        # Add separator after each state except the last
        if i < len(states) - 1:
            separator_cols.append(col)
            headers.append("")  # placeholder for separator column
            col += 1

    total_cols = col - 1  # last used column number

    # ── Row 2: State group labels ─────────────────────────────────────────
    STATE_COLORS = [
        "1A237E", "E65100", "1B5E20", "6A1B9A", "B71C1C",
        "004D40", "F57F17", "0D47A1", "880E4F", "33691E",
    ]
    for i, state in enumerate(states):
        start_col = state_col_starts[state]
        end_col = start_col + 2  # 3 columns per state
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        label_cell = ws.cell(row=2, column=start_col, value=f"── {state} ──")
        color = STATE_COLORS[i % len(STATE_COLORS)]
        label_cell.font = Font(bold=True, size=11, color=color)
        label_cell.alignment = CENTER_ALIGN

    # ── Row 3: Column headers ─────────────────────────────────────────────
    # Write headers but skip separator positions
    header_idx = 0
    for c in range(1, total_cols + 1):
        if c - 1 < len(headers):
            val = headers[c - 1]
            if val:  # skip empty separator placeholders
                cell = ws.cell(row=3, column=c, value=val)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER_ALIGN if c > 1 else LEFT_ALIGN
            elif c in separator_cols:
                # Separator column in header row — leave blank but apply header fill
                pass
        header_idx += 1

    # Also apply header fill to column A
    cell_a = ws.cell(row=3, column=1)
    if not cell_a.value:
        cell_a.value = "Line Item"
    cell_a.font = HEADER_FONT
    cell_a.fill = HEADER_FILL
    cell_a.alignment = LEFT_ALIGN

    # Merge title across all columns
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(total_cols, 30))

    # ── Write data rows ───────────────────────────────────────────────────
    separator_set = set(separator_cols)
    row_num = 4
    for label, row_type, is_revenue_like in PNL_STRUCTURE:
        is_pct_row = row_type == "pct_row"

        if row_type == "blank":
            row_num += 1
            continue

        ws.cell(row=row_num, column=1, value=label)

        if row_type not in ("header",):
            num_fmt = PCT_FMT if is_pct_row else DOLLAR_FMT
            var_fmt = PCT_FMT if is_pct_row else DOLLAR_VAR_FMT

            for state in states:
                col_offset = state_col_starts[state]
                state_rows = state_variances.get(state, [])
                # Find matching row
                match = None
                for sr in state_rows:
                    if sr["label"] == label:
                        match = sr
                        break

                if match and match["budget"] is not None:
                    # Actual (first)
                    c = ws.cell(row=row_num, column=col_offset, value=match["actual"])
                    c.number_format = num_fmt
                    c.alignment = RIGHT_ALIGN

                    # Budget (second)
                    c = ws.cell(row=row_num, column=col_offset + 1, value=match["budget"])
                    c.number_format = num_fmt
                    c.alignment = RIGHT_ALIGN

                    # $ Var
                    c = ws.cell(row=row_num, column=col_offset + 2, value=match["dollar_var"])
                    c.number_format = var_fmt
                    c.alignment = RIGHT_ALIGN
                    _apply_variance_color(c, match["favorable"])

        _format_row(ws, row_num, row_type if not is_pct_row else "item", 2, total_cols,
                    skip_cols=separator_set)
        row_num += 1

    last_row = row_num - 1

    # ── Apply separator column formatting ─────────────────────────────────
    for sep_col in separator_cols:
        _write_separator_col(ws, sep_col, last_row)
        ws.column_dimensions[get_column_letter(sep_col)].width = 2

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 30
    for c in range(2, total_cols + 1):
        if c not in separator_set:
            ws.column_dimensions[get_column_letter(c)].width = 14

    ws.freeze_panes = "B4"


def write_insights_sheet(wb, insights, month):
    """
    Write the Key Insights sheet.
    """
    ws = wb.create_sheet("Key Insights")

    # Title
    ws.cell(row=1, column=1, value=f"Treetop Therapy — Key Insights: {month} 2026")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1A237E")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # Headers
    headers = ["#", "Severity", "Category", "Insight", "$ Impact", "Action"]
    _write_header_row(ws, 3, headers)

    row_num = 4
    for i, insight in enumerate(insights, 1):
        sev = insight["severity"]

        ws.cell(row=row_num, column=1, value=i).alignment = CENTER_ALIGN

        c = ws.cell(row=row_num, column=2, value=sev.upper())
        c.font = SEVERITY_FONTS.get(sev, ITEM_FONT)
        c.fill = SEVERITY_FILLS.get(sev, PatternFill())
        c.alignment = CENTER_ALIGN

        ws.cell(row=row_num, column=3, value=insight["category"])

        ws.cell(row=row_num, column=4, value=insight["insight"]).font = ITEM_FONT

        c = ws.cell(row=row_num, column=5, value=insight.get("dollar_impact", 0))
        c.number_format = DOLLAR_VAR_FMT
        c.alignment = RIGHT_ALIGN

        ws.cell(row=row_num, column=6, value=insight.get("action", ""))

        # Light row border
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = THIN_BORDER

        row_num += 1

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 65
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 50

    ws.freeze_panes = "A4"


def write_waterfall_sheet(wb, waterfall, month):
    """
    Write the Variance Waterfall (EBITDA bridge) sheet.
    """
    ws = wb.create_sheet("EBITDA Waterfall")

    # Title
    ws.cell(row=1, column=1, value=f"Treetop Therapy — EBITDA Bridge: {month} 2026")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1A237E")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    # Headers
    headers = ["Category", "Impact ($)", "Running Total"]
    _write_header_row(ws, 3, headers)

    row_num = 4
    running = 0
    for i, (label, value) in enumerate(waterfall):
        ws.cell(row=row_num, column=1, value=label)

        c = ws.cell(row=row_num, column=2, value=value)
        c.number_format = DOLLAR_VAR_FMT
        c.alignment = RIGHT_ALIGN

        if i == 0:
            # Budget EBITDA — starting point
            running = value
            c.font = TOTAL_FONT
            ws.cell(row=row_num, column=1).font = TOTAL_FONT
        elif i == len(waterfall) - 1:
            # Actual EBITDA — ending point
            running = value
            c.font = TOTAL_FONT
            ws.cell(row=row_num, column=1).font = TOTAL_FONT
            for col in range(1, 4):
                ws.cell(row=row_num, column=col).border = THICK_BORDER
                ws.cell(row=row_num, column=col).fill = TOTAL_FILL
        else:
            # Variance items
            running += value
            if value > 0:
                c.font = GREEN_FONT
                c.fill = GREEN_FILL
            elif value < 0:
                c.font = RED_FONT
                c.fill = RED_FILL

        rt = ws.cell(row=row_num, column=3, value=running)
        rt.number_format = DOLLAR_FMT
        rt.alignment = RIGHT_ALIGN

        row_num += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    ws.freeze_panes = "A4"


def write_monthly_trends_sheet(wb, budget_data, actuals_by_month, months_loaded, working_days=None):
    """
    Write the Monthly Trends sheet.
    Shows key metrics month-by-month with improved formatting:
    - Only show months that have data (actual or budget)
    - Include Rev/Working Day and EBITDA/Working Day
    - Color-code actual vs budget variance
    - Group label row for Actuals and Budget
    """
    ws = wb.create_sheet("Monthly Trends")

    # Title
    ws.cell(row=1, column=1, value="Treetop Therapy — Monthly Trends: 2026")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1A237E")

    from config import MONTHS

    # Determine which months to show (only those with actual or budget data)
    active_months = []
    for m in MONTHS:
        has_actual = m in actuals_by_month and actuals_by_month[m].get("Total Revenue", 0)
        has_budget = False
        if "Total Revenue" in budget_data and m in budget_data["Total Revenue"]:
            has_budget = budget_data["Total Revenue"][m] != 0
        if has_actual or has_budget:
            active_months.append(m)

    # Cap at reasonable number — show months through current + a couple future
    if not active_months:
        active_months = MONTHS[:3]

    # Determine "current year" months (Jan-Dec 2026) vs historical (Oct-Dec 2025)
    # Historical months have actuals but their budget is for 2026 (not comparable)
    # We detect this by checking if historical months come AFTER the current month in MONTHS order
    current_year_months = set()
    for m in active_months:
        m_idx = MONTHS.index(m)
        # Months Jan (0) through Sep (8) are always current year
        # Months Oct (9) through Dec (11) are prior year IF they have actuals
        # but the current month is Jan (0) or early in the year
        # Simple heuristic: if month index > 8 (Oct+) and we have actual data,
        # it's likely prior year data; only compare to budget if it's current year
        if m_idx <= 8:
            current_year_months.add(m)
        elif m in actuals_by_month and actuals_by_month[m].get("Total Revenue", 0):
            # Has actuals — this is prior year historical, don't compare to budget
            pass
        else:
            current_year_months.add(m)  # Future budget-only month

    key_metrics = [
        ("REVENUE", "header"),
        ("Total Revenue", "total"),
        ("Rev / Working Day", "computed_rpd"),
        ("", "blank"),
        ("COST OF SERVICES", "header"),
        ("Total COGS", "subtotal"),
        ("Gross Profit", "total"),
        ("Gross Margin %", "pct_gm"),
        ("", "blank"),
        ("OPERATING EXPENSES", "header"),
        ("Billing Expense", "item"),
        ("Total Sales & Marketing", "subtotal"),
        ("Total StateOps Expense", "subtotal"),
        ("Total Clinic G&A", "subtotal"),
        ("Other Direct G&A Expense", "subtotal"),
        ("Total Corporate Expense", "subtotal"),
        ("Total Expenses", "subtotal"),
        ("", "blank"),
        ("EBITDA", "total"),
        ("EBITDA Margin %", "pct_ebitda"),
        ("EBITDA / Working Day", "computed_epd"),
    ]

    # Build headers: Metric | Oct Act | Oct Bud | ... | YTD Act | YTD Bud | Annual Bud
    # Row 2: group labels
    ACTUALS_LABEL_FILL = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
    BUDGET_LABEL_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    headers = ["Metric"]
    col = 2
    for m in active_months:
        has_act = m in actuals_by_month and actuals_by_month[m].get("Total Revenue", 0)
        is_current_year = m in current_year_months
        if has_act and is_current_year:
            headers.append(f"{m} Act")
            headers.append(f"{m} Bud")
            headers.append(f"{m} Var")
        elif has_act:
            # Historical month — actuals only, no budget comparison
            headers.append(f"{m} Act")
        else:
            headers.append(f"{m} Bud")

    # YTD Performance vs Plan columns — for current year months with actuals
    ytd_months = [m for m in active_months if m in current_year_months and m in actuals_by_month and actuals_by_month[m].get("Total Revenue", 0)]
    has_ytd = len(ytd_months) >= 1
    if has_ytd:
        headers.append("YTD Act")
        headers.append("YTD Bud")
        headers.append("YTD $ Var")
        headers.append("YTD % Var")

    headers.append("Annual Bud")

    _write_header_row(ws, 3, headers)
    total_cols = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(total_cols, 26))

    # Row 2: "Performance vs Plan" group label above YTD columns
    if has_ytd:
        # Calculate YTD start column: count header entries before YTD
        ytd_start_col = total_cols - 4  # 4 YTD cols + 1 Annual Bud = 5 from end, YTD starts at -4
        # Actually calculate from headers list
        ytd_col_start = headers.index("YTD Act") + 1  # +1 because headers is 0-indexed but columns are 1-indexed
        ytd_col_end = ytd_col_start + 3  # 4 columns: Act, Bud, $ Var, % Var
        YTD_LABEL_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        YTD_LABEL_FONT = Font(bold=True, size=10, color="1B5E20")
        ws.merge_cells(start_row=2, start_column=ytd_col_start, end_row=2, end_column=ytd_col_end)
        c = ws.cell(row=2, column=ytd_col_start, value="── Performance vs Plan YTD ──")
        c.font = YTD_LABEL_FONT
        c.fill = YTD_LABEL_FILL
        c.alignment = CENTER_ALIGN
        for cc in range(ytd_col_start, ytd_col_end + 1):
            ws.cell(row=2, column=cc).fill = YTD_LABEL_FILL

    # Write data rows
    row_num = 4
    for label, metric_type in key_metrics:
        if metric_type == "blank":
            row_num += 1
            continue

        ws.cell(row=row_num, column=1, value=label)

        if metric_type == "header":
            ws.cell(row=row_num, column=1).font = SECTION_FONT
            ws.cell(row=row_num, column=1).fill = SECTION_FILL
            for c in range(1, total_cols + 1):
                ws.cell(row=row_num, column=c).fill = SECTION_FILL
            row_num += 1
            continue

        # Determine formatting
        is_pct = metric_type in ("pct_gm", "pct_ebitda")
        is_computed = metric_type in ("computed_rpd", "computed_epd")
        is_total = metric_type == "total"
        is_subtotal = metric_type == "subtotal"

        font = TOTAL_FONT if is_total else (SUBTOTAL_FONT if is_subtotal else ITEM_FONT)
        if is_computed:
            font = Font(size=9, italic=True, color="555555")
        ws.cell(row=row_num, column=1).font = font
        if not is_total and not is_subtotal and metric_type != "header":
            ws.cell(row=row_num, column=1).alignment = INDENT_ALIGN

        col = 2
        ytd_act = 0
        ytd_bud = 0
        annual_bud = 0

        for m in active_months:
            has_act = m in actuals_by_month and actuals_by_month[m].get("Total Revenue", 0)
            is_cy = m in current_year_months
            # Column mode: "abv" = act+bud+var, "a" = act only, "b" = bud only
            col_mode = "abv" if (has_act and is_cy) else ("a" if has_act else "b")

            if is_pct:
                # Compute pct values
                act_val = None
                bud_val = None
                if has_act:
                    rev_act = actuals_by_month[m].get("Total Revenue", 0)
                    if metric_type == "pct_gm":
                        act_val = actuals_by_month[m].get("Gross Profit", 0) / rev_act if rev_act else 0
                    else:
                        act_val = actuals_by_month[m].get("EBITDA", 0) / rev_act if rev_act else 0
                if is_cy:
                    rev_bud = budget_data.get("Total Revenue", {}).get(m, 0)
                    if metric_type == "pct_gm":
                        bud_val = budget_data.get("Gross Profit", {}).get(m, 0) / rev_bud if rev_bud else 0
                    else:
                        bud_val = budget_data.get("EBITDA", {}).get(m, 0) / rev_bud if rev_bud else 0

                if col_mode == "abv":
                    if act_val is not None:
                        c = ws.cell(row=row_num, column=col, value=act_val)
                        c.number_format = PCT_FMT; c.alignment = RIGHT_ALIGN; c.font = font
                    col += 1
                    if bud_val is not None:
                        c = ws.cell(row=row_num, column=col, value=bud_val)
                        c.number_format = PCT_FMT; c.alignment = RIGHT_ALIGN
                    col += 1
                    if act_val is not None and bud_val is not None:
                        pp = act_val - bud_val
                        c = ws.cell(row=row_num, column=col, value=pp)
                        c.number_format = PCT_FMT; c.alignment = RIGHT_ALIGN
                        _apply_variance_color(c, pp >= 0)
                    col += 1
                elif col_mode == "a":
                    if act_val is not None:
                        c = ws.cell(row=row_num, column=col, value=act_val)
                        c.number_format = PCT_FMT; c.alignment = RIGHT_ALIGN; c.font = font
                    col += 1
                else:
                    if bud_val is not None:
                        c = ws.cell(row=row_num, column=col, value=bud_val)
                        c.number_format = PCT_FMT; c.alignment = RIGHT_ALIGN
                    col += 1

            elif is_computed:
                wd = (working_days or {}).get(m)
                act_base = None
                bud_base = None
                if has_act and wd:
                    act_base = actuals_by_month[m].get("Total Revenue" if metric_type == "computed_rpd" else "EBITDA", 0)
                if is_cy and wd:
                    bud_base = budget_data.get("Total Revenue" if metric_type == "computed_rpd" else "EBITDA", {}).get(m, 0)

                if col_mode == "abv" and wd:
                    act_v = act_base / wd if act_base else 0
                    bud_v = bud_base / wd if bud_base else 0
                    c = ws.cell(row=row_num, column=col, value=act_v)
                    c.number_format = DOLLAR_FMT; c.alignment = RIGHT_ALIGN; c.font = font
                    col += 1
                    c = ws.cell(row=row_num, column=col, value=bud_v)
                    c.number_format = DOLLAR_FMT; c.alignment = RIGHT_ALIGN
                    col += 1
                    var_v = act_v - bud_v
                    c = ws.cell(row=row_num, column=col, value=var_v)
                    c.number_format = DOLLAR_VAR_FMT; c.alignment = RIGHT_ALIGN
                    _apply_variance_color(c, var_v >= 0)
                    col += 1
                elif col_mode == "abv":
                    col += 3
                elif col_mode == "a" and wd and act_base:
                    c = ws.cell(row=row_num, column=col, value=act_base / wd)
                    c.number_format = DOLLAR_FMT; c.alignment = RIGHT_ALIGN; c.font = font
                    col += 1
                elif col_mode == "a":
                    col += 1
                elif col_mode == "b" and wd and bud_base:
                    c = ws.cell(row=row_num, column=col, value=bud_base / wd)
                    c.number_format = DOLLAR_FMT; c.alignment = RIGHT_ALIGN
                    col += 1
                else:
                    col += 1
            else:
                # Standard dollar metric
                act_val = actuals_by_month[m].get(label, 0) if has_act else 0
                bud_val = budget_data.get(label, {}).get(m, 0) if is_cy else 0
                annual_bud += budget_data.get(label, {}).get(m, 0)

                if col_mode == "abv":
                    c = ws.cell(row=row_num, column=col, value=act_val)
                    c.number_format = DOLLAR_FMT; c.alignment = RIGHT_ALIGN; c.font = font
                    ytd_act += act_val
                    col += 1
                    c = ws.cell(row=row_num, column=col, value=bud_val)
                    c.number_format = DOLLAR_FMT; c.alignment = RIGHT_ALIGN
                    ytd_bud += bud_val
                    col += 1
                    var_val = act_val - bud_val
                    c = ws.cell(row=row_num, column=col, value=var_val)
                    c.number_format = DOLLAR_VAR_FMT; c.alignment = RIGHT_ALIGN
                    is_rev_like = label in ("Total Revenue", "Gross Profit", "EBITDA")
                    _apply_variance_color(c, var_val >= 0 if is_rev_like else var_val <= 0)
                    col += 1
                elif col_mode == "a":
                    c = ws.cell(row=row_num, column=col, value=act_val)
                    c.number_format = DOLLAR_FMT; c.alignment = RIGHT_ALIGN; c.font = font
                    ytd_act += act_val
                    col += 1
                else:
                    c = ws.cell(row=row_num, column=col, value=bud_val)
                    c.number_format = DOLLAR_FMT; c.alignment = RIGHT_ALIGN
                    col += 1

        # YTD Performance vs Plan columns
        if has_ytd and not is_pct and not is_computed:
            c = ws.cell(row=row_num, column=col, value=ytd_act)
            c.number_format = DOLLAR_FMT
            c.alignment = RIGHT_ALIGN
            c.font = SUBTOTAL_FONT
            col += 1
            c = ws.cell(row=row_num, column=col, value=ytd_bud)
            c.number_format = DOLLAR_FMT
            c.alignment = RIGHT_ALIGN
            col += 1
            ytd_var = ytd_act - ytd_bud
            c = ws.cell(row=row_num, column=col, value=ytd_var)
            c.number_format = DOLLAR_VAR_FMT
            c.alignment = RIGHT_ALIGN
            is_rev = label in ("Total Revenue", "Gross Profit", "EBITDA",
                               "Gross Profit Net Billing")
            _apply_variance_color(c, ytd_var >= 0 if is_rev else ytd_var <= 0)
            col += 1
            # YTD % Variance
            if ytd_bud != 0:
                ytd_pct = ytd_var / abs(ytd_bud)
                c = ws.cell(row=row_num, column=col, value=ytd_pct)
                c.number_format = PCT_FMT
                c.alignment = RIGHT_ALIGN
                _apply_variance_color(c, ytd_var >= 0 if is_rev else ytd_var <= 0)
            col += 1
        elif has_ytd and is_pct:
            # For pct rows: show YTD actual %, YTD budget %, pp diff, skip % var
            ytd_rev_act = sum(
                actuals_by_month[m].get("Total Revenue", 0)
                for m in ytd_months
            )
            ytd_rev_bud = sum(
                budget_data.get("Total Revenue", {}).get(m, 0)
                for m in ytd_months
            )
            if metric_type == "pct_gm":
                ytd_num_act = sum(actuals_by_month[m].get("Gross Profit", 0) for m in ytd_months)
                ytd_num_bud = sum(budget_data.get("Gross Profit", {}).get(m, 0) for m in ytd_months)
            else:  # pct_ebitda
                ytd_num_act = sum(actuals_by_month[m].get("EBITDA", 0) for m in ytd_months)
                ytd_num_bud = sum(budget_data.get("EBITDA", {}).get(m, 0) for m in ytd_months)

            ytd_pct_act = ytd_num_act / ytd_rev_act if ytd_rev_act else 0
            ytd_pct_bud = ytd_num_bud / ytd_rev_bud if ytd_rev_bud else 0
            pp_diff = ytd_pct_act - ytd_pct_bud

            c = ws.cell(row=row_num, column=col, value=ytd_pct_act)
            c.number_format = PCT_FMT; c.alignment = RIGHT_ALIGN; c.font = font
            col += 1
            c = ws.cell(row=row_num, column=col, value=ytd_pct_bud)
            c.number_format = PCT_FMT; c.alignment = RIGHT_ALIGN
            col += 1
            c = ws.cell(row=row_num, column=col, value=pp_diff)
            c.number_format = PCT_FMT; c.alignment = RIGHT_ALIGN
            _apply_variance_color(c, pp_diff >= 0)
            col += 1
            col += 1  # Skip % Var column for pct rows
        elif has_ytd and is_computed:
            col += 4  # Skip YTD columns for computed rows

        # Annual budget
        if not is_pct and not is_computed:
            # Compute annual budget from all 12 months
            annual_total = sum(budget_data.get(label, {}).get(m, 0) for m in MONTHS)
            c = ws.cell(row=row_num, column=col, value=annual_total)
            c.number_format = DOLLAR_FMT
            c.alignment = RIGHT_ALIGN
            c.font = SUBTOTAL_FONT

        # Row borders for totals
        if is_total:
            for cc in range(1, total_cols + 1):
                ws.cell(row=row_num, column=cc).font = TOTAL_FONT
                ws.cell(row=row_num, column=cc).border = THIN_BORDER

        row_num += 1

    # Working days reference row
    row_num += 1
    ws.cell(row=row_num, column=1, value="Working Days").font = Font(size=9, italic=True, color="666666")
    col = 2
    for m in active_months:
        has_act = m in actuals_by_month and actuals_by_month[m].get("Total Revenue", 0)
        is_cy = m in current_year_months
        wd = (working_days or {}).get(m, "")
        col_mode = "abv" if (has_act and is_cy) else ("a" if has_act else "b")
        ws.cell(row=row_num, column=col, value=wd).font = Font(size=9, italic=True, color="666666")
        ws.cell(row=row_num, column=col).alignment = CENTER_ALIGN
        col += 3 if col_mode == "abv" else 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    for i in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13

    ws.freeze_panes = "B4"


def write_clinic_detail_sheet(wb, clinics_detail, month):
    """
    Write the Clinic-by-Clinic detail tab with:
    - Full P&L detail per clinic (not summarized)
    - Clinics grouped by state with state subtotals
    - Gross Margin % and EBITDA % inline
    - Grand total column
    """
    ws = wb.create_sheet("Clinic Detail")

    ws.cell(row=1, column=1, value=f"Treetop Therapy — Clinic-by-Clinic Detail: {month} 2026")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1A237E")

    if not clinics_detail:
        ws.cell(row=3, column=1, value="No individual clinic data available.")
        return

    # Group clinics by state (prefix before the dash)
    from collections import OrderedDict
    state_groups = OrderedDict()
    for name in sorted(clinics_detail.keys()):
        state = name.split("-")[0] if "-" in name else name
        if state not in state_groups:
            state_groups[state] = []
        state_groups[state].append(name)

    # Build ordered column list with separator columns between state groups
    # col_entries: list of (label, type) where type = "clinic", "state_total", "grand_total", "separator"
    col_entries = []
    state_list = list(state_groups.keys())
    for si, state in enumerate(state_list):
        clinics = state_groups[state]
        for clinic_name in clinics:
            col_entries.append((clinic_name, "clinic"))
        if len(clinics) > 1:
            col_entries.append((f"{state} Total", "state_total"))
        # Add separator between state groups (not after last state before Grand Total)
        if si < len(state_list) - 1:
            col_entries.append(("", "separator"))
    # Separator before Grand Total
    col_entries.append(("", "separator"))
    col_entries.append(("Grand Total", "grand_total"))

    # Build column mapping and track separator positions
    separator_cols = set()
    col_map = {}  # index in col_entries -> actual column number
    col = 2
    for idx, (entry_label, entry_type) in enumerate(col_entries):
        col_map[idx] = col
        if entry_type == "separator":
            separator_cols.add(col)
        col += 1
    total_cols = col - 1

    # Full P&L structure for clinic detail (with inline computed rows)
    from config import PNL_STRUCTURE
    metrics = []
    for label, row_type, is_rev in PNL_STRUCTURE:
        # Skip Corporate section (clinics don't have corporate expenses)
        if label in ("CORPORATE", "Corporate Overhead Wages", "Corporate Overhead Bonus",
                     "Corporate Rent", "Corporate Office Expense", "Total Corporate Expense"):
            continue
        metrics.append((label, row_type, is_rev))

    # Row 2: State group labels with colors
    STATE_COLORS = [
        "1A237E", "E65100", "1B5E20", "6A1B9A", "B71C1C",
        "004D40", "F57F17", "0D47A1", "880E4F", "33691E",
    ]
    STATE_GROUP_FILL = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")

    for si, state in enumerate(state_list):
        clinics = state_groups[state]
        num_data_cols = len(clinics) + (1 if len(clinics) > 1 else 0)
        # Find start column for this state's clinics
        state_start_col = None
        for idx, (entry_label, entry_type) in enumerate(col_entries):
            if entry_type == "clinic" and entry_label in clinics and state_start_col is None:
                state_start_col = col_map[idx]
                break
        if state_start_col is None:
            continue
        end_col = state_start_col + num_data_cols - 1

        if num_data_cols > 1:
            ws.merge_cells(start_row=2, start_column=state_start_col, end_row=2, end_column=end_col)
        color = STATE_COLORS[si % len(STATE_COLORS)]
        c = ws.cell(row=2, column=state_start_col, value=f"── {state} ──")
        c.font = Font(bold=True, size=10, color=color)
        c.fill = STATE_GROUP_FILL
        c.alignment = CENTER_ALIGN
        for cc in range(state_start_col, end_col + 1):
            ws.cell(row=2, column=cc).fill = STATE_GROUP_FILL

    # Headers (row 3)
    for idx, (entry_label, entry_type) in enumerate(col_entries):
        c_num = col_map[idx]
        if entry_type == "separator":
            continue  # skip separators in header
        cell = ws.cell(row=3, column=c_num, value=entry_label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
    # Column A header
    cell_a = ws.cell(row=3, column=1, value="Line Item")
    cell_a.font = HEADER_FONT
    cell_a.fill = HEADER_FILL
    cell_a.alignment = LEFT_ALIGN

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(total_cols, 30))

    # Precompute state totals
    state_totals = {}
    for state, clinics in state_groups.items():
        state_data = {}
        for clinic_name in clinics:
            cdata = clinics_detail.get(clinic_name, {})
            for item, val in cdata.items():
                state_data[item] = state_data.get(item, 0) + (val or 0)
        state_totals[state] = state_data

    # Grand total
    grand_total = {}
    for cdata in clinics_detail.values():
        for item, val in cdata.items():
            grand_total[item] = grand_total.get(item, 0) + (val or 0)

    # Write data rows
    row_num = 4
    for label, row_type, is_rev in metrics:
        is_pct_row = row_type == "pct_row"

        if row_type == "blank":
            row_num += 1
            continue

        ws.cell(row=row_num, column=1, value=label)

        if row_type not in ("header",):
            for idx, (entry_label, entry_type) in enumerate(col_entries):
                if entry_type == "separator":
                    continue  # skip separator columns

                col = col_map[idx]

                # Get value based on entry type
                if entry_type == "clinic":
                    source = clinics_detail.get(entry_label, {})
                elif entry_type == "state_total":
                    state_name = entry_label.replace(" Total", "")
                    source = state_totals.get(state_name, {})
                else:  # grand_total
                    source = grand_total

                if is_pct_row:
                    # Compute percentage inline
                    rev = source.get("Total Revenue", 0)
                    if rev and rev > 0:
                        if "Gross Margin" in label and "Net" not in label:
                            gp = source.get("Gross Profit", 0)
                            val = gp / rev if gp else 0
                        elif "Billing Expense" in label:
                            be = source.get("Billing Expense", 0)
                            val = be / rev if be else 0
                        elif "Net Billing" in label:
                            gpnb = source.get("Gross Profit Net Billing", 0)
                            val = gpnb / rev if gpnb else 0
                        else:
                            val = 0
                        if val:
                            c = ws.cell(row=row_num, column=col, value=val)
                            c.number_format = PCT_FMT
                            c.alignment = RIGHT_ALIGN
                else:
                    val = source.get(label, 0)
                    if val:
                        c = ws.cell(row=row_num, column=col, value=val)
                        c.number_format = DOLLAR_FMT
                        c.alignment = RIGHT_ALIGN
                        # Bold for state and grand totals
                        if entry_type in ("state_total", "grand_total"):
                            c.font = SUBTOTAL_FONT

        _format_row(ws, row_num, row_type if not is_pct_row else "item", 2, total_cols,
                    skip_cols=separator_cols)
        row_num += 1

    # Add EBITDA % row at the end
    row_num += 1
    ws.cell(row=row_num, column=1, value="EBITDA Margin %").font = SUBTOTAL_FONT
    for idx, (entry_label, entry_type) in enumerate(col_entries):
        if entry_type == "separator":
            continue

        col = col_map[idx]

        if entry_type == "clinic":
            source = clinics_detail.get(entry_label, {})
        elif entry_type == "state_total":
            state_name = entry_label.replace(" Total", "")
            source = state_totals.get(state_name, {})
        else:
            source = grand_total

        rev = source.get("Total Revenue", 0)
        ebitda = source.get("EBITDA", 0)
        if rev and rev > 0 and ebitda:
            c = ws.cell(row=row_num, column=col, value=ebitda / rev)
            c.number_format = PCT_FMT
            c.alignment = RIGHT_ALIGN
            if entry_type in ("state_total", "grand_total"):
                c.font = SUBTOTAL_FONT
            # Color code: green if positive, red if negative
            _apply_variance_color(c, ebitda >= 0)

    last_row = row_num

    # Apply separator column formatting
    for sep_col in separator_cols:
        _write_separator_col(ws, sep_col, last_row)
        ws.column_dimensions[get_column_letter(sep_col)].width = 2

    ws.column_dimensions["A"].width = 28
    for i in range(2, total_cols + 1):
        if i not in separator_cols:
            ws.column_dimensions[get_column_letter(i)].width = 15
    ws.freeze_panes = "B4"


def write_management_sheet(wb, mgmt_actuals, mgmt_budget, month):
    """
    Write the Management/Corporate budget vs actual tab.
    """
    ws = wb.create_sheet("Management BvA")

    ws.cell(row=1, column=1, value=f"Treetop Therapy — Management/Corporate: {month} 2026")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1A237E")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    headers = ["Line Item", f"{month} Actual", f"{month} Budget", "$ Variance", "% Variance"]
    _write_header_row(ws, 3, headers)

    from config import PNL_STRUCTURE

    row_num = 4
    for label, row_type, is_revenue_like in PNL_STRUCTURE:
        if row_type in ("blank", "pct_row"):
            if row_type == "blank":
                row_num += 1
            continue

        ws.cell(row=row_num, column=1, value=label)

        if row_type != "header":
            act_val = mgmt_actuals.get(label, 0) if mgmt_actuals else 0
            # Budget for MGMT: corporate items from the WholeCo budget
            bud_val = mgmt_budget.get(label, {}).get(month, 0) if isinstance(mgmt_budget.get(label), dict) else mgmt_budget.get(label, 0)

            if act_val or bud_val:
                c = ws.cell(row=row_num, column=2, value=act_val)
                c.number_format = DOLLAR_FMT
                c.alignment = RIGHT_ALIGN

                c = ws.cell(row=row_num, column=3, value=bud_val)
                c.number_format = DOLLAR_FMT
                c.alignment = RIGHT_ALIGN

                dollar_var = act_val - bud_val
                c = ws.cell(row=row_num, column=4, value=dollar_var)
                c.number_format = DOLLAR_VAR_FMT
                c.alignment = RIGHT_ALIGN
                if is_revenue_like is not None:
                    fav = dollar_var >= 0 if is_revenue_like else dollar_var <= 0
                    _apply_variance_color(c, fav)

                if bud_val != 0:
                    pct_var = dollar_var / abs(bud_val)
                    c = ws.cell(row=row_num, column=5, value=pct_var)
                    c.number_format = PCT_FMT
                    c.alignment = RIGHT_ALIGN
                    if is_revenue_like is not None:
                        _apply_variance_color(c, fav)

        _format_row(ws, row_num, row_type, 2, 5)
        row_num += 1

    ws.column_dimensions["A"].width = 30
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["E"].width = 14
    ws.freeze_panes = "B4"


def write_margin_analysis_sheet(wb, margin_analysis, month):
    """
    Write the Gross Margin Analysis (driver decomposition) sheet.
    """
    ws = wb.create_sheet("Margin Analysis")

    ws.cell(row=1, column=1, value=f"Treetop Therapy — Gross Margin Variance Analysis: {month} 2026")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1A237E")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    headers = ["Category", "Metric", "Actual", "Budget", "Variance", "Insight"]
    _write_header_row(ws, 3, headers)

    row_num = 4
    prev_category = None
    for item in margin_analysis:
        cat = item["category"]
        metric = item["metric"]

        # Section separator
        if cat != prev_category:
            if prev_category is not None:
                row_num += 1
            ws.cell(row=row_num, column=1, value=cat).font = SECTION_FONT
            ws.cell(row=row_num, column=1).fill = SECTION_FILL
            for col in range(1, 7):
                ws.cell(row=row_num, column=col).fill = SECTION_FILL
            row_num += 1
            prev_category = cat

        ws.cell(row=row_num, column=1, value="")
        ws.cell(row=row_num, column=2, value=metric).font = ITEM_FONT

        is_pct = "%" in metric or "Share" in metric or "Margin" in metric
        fmt = PCT_FMT if is_pct else DOLLAR_FMT
        var_fmt = PCT_FMT if is_pct else DOLLAR_VAR_FMT

        c = ws.cell(row=row_num, column=3, value=item["actual_val"])
        c.number_format = fmt
        c.alignment = RIGHT_ALIGN

        c = ws.cell(row=row_num, column=4, value=item["budget_val"])
        c.number_format = fmt
        c.alignment = RIGHT_ALIGN

        var = item["variance"]
        c = ws.cell(row=row_num, column=5, value=var)
        c.number_format = var_fmt
        c.alignment = RIGHT_ALIGN
        if var != 0:
            # For revenue/profit metrics: positive = green
            # For cost metrics: negative = green
            if any(kw in metric for kw in ["Revenue", "Profit", "Margin"]):
                _apply_variance_color(c, var >= 0)
            else:
                _apply_variance_color(c, var <= 0)

        ws.cell(row=row_num, column=6, value=item["insight"]).font = ITEM_FONT
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = THIN_BORDER
        row_num += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 75
    ws.freeze_panes = "A4"


def write_data_quality_sheet(wb, issues, month):
    """
    Write the Data Quality / Anomaly Detection sheet.
    """
    ws = wb.create_sheet("Data Quality")

    ws.cell(row=1, column=1, value=f"Treetop Therapy — Data Quality Checks: {month} 2026")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1A237E")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    if not issues:
        ws.cell(row=3, column=1, value="All data quality checks passed. No anomalies detected.")
        ws.cell(row=3, column=1).font = Font(size=11, color="1B5E20", bold=True)
        return

    # Summary
    critical_count = sum(1 for i in issues if i["severity"] == "critical")
    warning_count = sum(1 for i in issues if i["severity"] == "warning")
    ws.cell(row=2, column=1, value=f"Found {len(issues)} issues: {critical_count} critical, {warning_count} warnings")
    ws.cell(row=2, column=1).font = Font(size=10, italic=True, color="666666")

    headers = ["#", "Severity", "Category", "Issue", "Detail"]
    _write_header_row(ws, 4, headers)

    row_num = 5
    for i, issue in enumerate(issues, 1):
        sev = issue["severity"]
        ws.cell(row=row_num, column=1, value=i).alignment = CENTER_ALIGN

        c = ws.cell(row=row_num, column=2, value=sev.upper())
        c.font = SEVERITY_FONTS.get(sev, ITEM_FONT)
        c.fill = SEVERITY_FILLS.get(sev, PatternFill())
        c.alignment = CENTER_ALIGN

        ws.cell(row=row_num, column=3, value=issue["category"])
        ws.cell(row=row_num, column=4, value=issue["issue"]).font = ITEM_FONT
        ws.cell(row=row_num, column=5, value=issue["detail"]).font = Font(size=9, color="555555")

        for col in range(1, 6):
            ws.cell(row=row_num, column=col).border = THIN_BORDER
        row_num += 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 70
    ws.freeze_panes = "A5"


def write_unmapped_sheet(wb, unmapped_transactions):
    """
    Write the Unmapped Transactions sheet.
    Section 1: Summary by GL account (name, count, total $)
    Section 2: Full transaction detail
    """
    ws = wb.create_sheet("Unmapped Transactions")

    # Title
    c = ws.cell(row=1, column=1, value="Unmapped Transactions")
    c.font = Font(bold=True, size=14)

    c = ws.cell(row=2, column=1,
                value="These GL accounts could not be mapped to a P&L line item. "
                      "Update the Mapping Tab to add new entries.")
    c.font = Font(italic=True, size=10, color="666666")

    # ── Section 1: Summary by GL Account ────────────────────────────────
    from collections import Counter, defaultdict
    acct_counts = Counter()
    acct_totals = defaultdict(float)
    for txn in unmapped_transactions:
        acct_counts[txn["account"]] += 1
        acct_totals[txn["account"]] += txn["amount"]

    row = 4
    summary_headers = ["GL Account", "# Transactions", "Total Amount"]
    for i, h in enumerate(summary_headers):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN if i > 0 else LEFT_ALIGN

    row = 5
    for acct, cnt in acct_counts.most_common():
        ws.cell(row=row, column=1, value=acct).font = ITEM_FONT
        ws.cell(row=row, column=2, value=cnt).alignment = CENTER_ALIGN
        c = ws.cell(row=row, column=3, value=acct_totals[acct])
        c.number_format = DOLLAR_FMT
        c.alignment = RIGHT_ALIGN
        row += 1

    # Total row
    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=row, column=2, value=len(unmapped_transactions)).alignment = CENTER_ALIGN
    c = ws.cell(row=row, column=3, value=sum(acct_totals.values()))
    c.number_format = DOLLAR_FMT
    c.font = TOTAL_FONT
    c.alignment = RIGHT_ALIGN

    # ── Section 2: Transaction Detail ───────────────────────────────────
    row += 3
    c = ws.cell(row=row, column=1, value="Transaction Detail")
    c.font = Font(bold=True, size=12)
    row += 1

    detail_headers = ["Date", "GL Account", "Item Class", "Amount",
                      "Transaction Type", "Full Name"]
    for i, h in enumerate(detail_headers):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN if i > 0 else LEFT_ALIGN

    row += 1
    for txn in sorted(unmapped_transactions, key=lambda t: t["account"]):
        ws.cell(row=row, column=1, value=txn["date"])
        ws.cell(row=row, column=2, value=txn["account"])
        ws.cell(row=row, column=3, value=txn["item_class"])
        c = ws.cell(row=row, column=4, value=txn["amount"])
        c.number_format = DOLLAR_FMT
        c.alignment = RIGHT_ALIGN
        ws.cell(row=row, column=5, value=txn["transaction_type"])
        ws.cell(row=row, column=6, value=txn["full_name"])
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 40
    ws.freeze_panes = "A5"


def build_output_workbook(
    wholeco_variance,
    segment_variance,
    state_variances,
    waterfall,
    insights,
    budget_data,
    actuals_by_month,
    months_loaded,
    month,
    states,
    output_path,
    prior_month_actuals=None,
    working_days=None,
    clinics_detail=None,
    mgmt_actuals=None,
    mgmt_budget=None,
    margin_analysis=None,
    data_quality_issues=None,
    unmapped_transactions=None,
):
    """
    Build the complete output workbook and save to disk.
    """
    wb = openpyxl.Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    # Sheet 1: WholeCo Summary (with MoM and Rev/Working Day)
    write_wholeco_sheet(wb, wholeco_variance, month,
                        prior_month_actuals=prior_month_actuals,
                        working_days=working_days)

    # Sheet 2: Home vs Clinic
    write_segment_sheet(wb, segment_variance, month)

    # Sheet 3: State Drill-Down
    write_state_drilldown_sheet(wb, state_variances, month, states)

    # Sheet 4: Clinic Detail
    if clinics_detail:
        write_clinic_detail_sheet(wb, clinics_detail, month)

    # Sheet 5: Monthly Trends (Management BvA removed per user request)
    write_monthly_trends_sheet(wb, budget_data, actuals_by_month, months_loaded,
                               working_days=working_days)

    # Sheet 7: Margin Analysis
    if margin_analysis:
        write_margin_analysis_sheet(wb, margin_analysis, month)

    # Sheet 8: Key Insights
    write_insights_sheet(wb, insights, month)

    # Sheet 9: EBITDA Waterfall
    write_waterfall_sheet(wb, waterfall, month)

    # Sheet 10: Data Quality
    if data_quality_issues is not None:
        write_data_quality_sheet(wb, data_quality_issues, month)

    # Sheet 11: Unmapped Transactions (only when using raw data parser)
    if unmapped_transactions:
        write_unmapped_sheet(wb, unmapped_transactions)

    wb.save(output_path)
    return output_path
